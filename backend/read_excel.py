import openpyxl
import sys

def print_excel(path):
    print(f'--- FILE: {path} ---')
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f'SHEET: {sheet_name}')
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    # Join with | to make it easier to read and ensure it prints to stdout
                    print(" | ".join([str(c) if c is not None else "" for c in row]))
    except Exception as e:
        print(f'Error reading {path}: {e}')

v3_path = '/Users/gabrielboian/Desktop/dev.nosync/t2c/alz-nota-futura/v3/Rascunho Banco de Dados - NF Entrega Futura1.xlsx'
v2_path = '/Users/gabrielboian/Desktop/dev.nosync/t2c/alz-nota-futura/v2/Rascunho Banco de Dados - NF Entrega Futura (1).xlsx'

print_excel(v3_path)
print_excel(v2_path)
