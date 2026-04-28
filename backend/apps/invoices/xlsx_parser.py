"""XLSX parser for the "Base de notas filhas" bulk upload.

Expected columns (header row detected automatically, case-insensitive,
accents stripped, order-agnostic):

    TE | TN | Tipo NF | Quantidade | Valor Bruto (NF) | Nr. Nota |
    Pedido SAP | Data Emissão | Data Movimento | DOC SAP (Contábil) |
    COD SAP (Cód Fornecedor) | Produtor (Emissor NF) | Produto |
    Tipo Produto | Filial | Contrato (Lote de Compra) |
    Produto (Emissor Contrato) | Safra |
    Funrural | Senar | Fethab | Facs | Fundeinfra | IMA

``Tipo NF`` drives the row classification:
    * ``Compra Futura``  → mother NF (``NFFutureDelivery``)
    * ``Remessa``        → child NF (``ChildNF``)

Rows are linked by ``Contrato (Lote de Compra)``: every Remessa row is a
child of the Compra Futura row sharing the same contract number.

See docs/05-workflows.md §FLOW 7 and docs/02-pages-and-ui.md §PAGE 13.
"""
from __future__ import annotations

import datetime as dt
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import IO, Any

import openpyxl


ROW_TYPE_MOTHER = 'mother'
ROW_TYPE_CHILD = 'child'


# Normalised header -> canonical field name.
# `_norm` lower-cases, NFKD-normalises (so "º" → "o"), strips combining
# marks and non-alphanumerics. So "Nº NF" → "no nf", "Nr. Nota" → "nr nota".
HEADER_MAP: dict[str, str] = {
    # Row type
    'tipo nf': 'row_type',
    # NF number (child NF number for Remessa rows; mother NF number for
    # Compra Futura rows).
    'nr nota': 'nf_number',
    'no nf': 'nf_number',
    'numero nf': 'nf_number',
    'numero nota': 'nf_number',
    # Reference numbers (TE/TN carry extra series info — optional).
    'te': 'te',
    'tn': 'tn',
    # Quantity (kg) — positive for mother, negative for child. Parser keeps
    # the absolute value on the child.
    'quantidade': 'quantity_kg',
    'qtd nf': 'quantity_kg',
    # Monetary
    'valor bruto nf': 'gross_value',
    'valor bruto': 'gross_value',
    'vlr bruto': 'gross_value',
    'vlr unitario': 'unit_value',
    'valor unitario': 'unit_value',
    # Dates
    'data emissao': 'issue_date',
    'data movimento': 'movement_date',
    # SAP identifiers
    'pedido sap': 'sap_order',
    'doc sap contabil': 'sap_doc',
    'cod sap cod fornecedor': 'sap_code',
    'cod sap': 'sap_code',
    'codigo sap': 'sap_code',
    # Producer / product / branch
    'produtor emissor nf': 'producer_name',
    'produtor': 'producer_name',
    'nome produtor': 'producer_name',
    'produto': 'product',
    'tipo produto': 'product_type',
    'filial': 'branch_name',
    # Contract lot — primary link between mother and children.
    'contrato lote de compra': 'lot_number',
    'contrato': 'lot_number',
    'no lote': 'lot_number',
    'numero lote': 'lot_number',
    # Producer that emitted the contract (optional)
    'produto emissor contrato': 'contract_producer_name',
    # Harvest year
    'safra': 'harvest_year',
    # Fiscal rates (optional, stored loosely)
    'funrural': 'funrural',
    'senar': 'senar',
    'fethab': 'fethab',
    'facs': 'facs',
    'fundeinfra': 'fundeinfra',
    'ima': 'ima',
    # State registration (kept for backwards compat with older spreadsheets)
    'inscricao estadual': 'state_registration',
    'ie': 'state_registration',
    # NF key (optional)
    'chave': 'nf_key',
    'chave nf': 'nf_key',
}

ROW_TYPE_MAP: dict[str, str] = {
    'compra futura': ROW_TYPE_MOTHER,
    'nf mae': ROW_TYPE_MOTHER,
    'nota mae': ROW_TYPE_MOTHER,
    'mae': ROW_TYPE_MOTHER,
    'remessa': ROW_TYPE_CHILD,
    'nf filha': ROW_TYPE_CHILD,
    'nota filha': ROW_TYPE_CHILD,
    'filha': ROW_TYPE_CHILD,
}

DECIMAL_FIELDS = {
    'quantity_kg', 'unit_value', 'gross_value',
    'funrural', 'senar', 'fethab', 'facs', 'fundeinfra', 'ima',
}


def _norm(s: Any) -> str:
    if s is None:
        return ''
    text = str(s).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    out = []
    for ch in text:
        if ch.isalnum() or ch == ' ':
            out.append(ch)
        else:
            out.append(' ')
    return ' '.join(''.join(out).split())


def _to_str(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return None
    s = str(v).strip()
    if ',' in s and '.' in s and s.rfind(',') > s.rfind('.'):
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace(',', '.')
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _to_date(v: Any):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str) and v.strip():
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return dt.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


@dataclass
class ParsedRow:
    row_number: int
    row_type: str  # ROW_TYPE_MOTHER or ROW_TYPE_CHILD
    data: dict[str, Any]
    error: str = ''


@dataclass
class ParseReport:
    rows_total: int = 0
    rows_valid: int = 0
    rows_invalid: int = 0
    parsed: list[ParsedRow] = field(default_factory=list)
    header_errors: list[str] = field(default_factory=list)


def _find_header(ws) -> tuple[int, dict[int, str]] | None:
    """Find the first row containing recognised header cells."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        mapping: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            key = _norm(cell)
            if not key:
                continue
            if key in HEADER_MAP:
                mapping[col_idx] = HEADER_MAP[key]
        values = set(mapping.values())
        if 'nf_number' in values and 'quantity_kg' in values:
            return i + 1, mapping
    return None


def parse_nf_xlsx(fp: IO[bytes]) -> ParseReport:
    """Parse a "Base de notas filhas" xlsx into classified rows.

    The caller (view) is responsible for persisting rows and linking
    each child to its mother via ``lot_number``.
    """
    report = ParseReport()
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    header = _find_header(ws)
    if header is None:
        report.header_errors.append(
            'Cabeçalho não encontrado. A planilha precisa conter pelo menos '
            '"Tipo NF", "Nr. Nota" e "Quantidade".'
        )
        return report

    header_row, col_map = header
    # When the spreadsheet has no "Tipo NF" column (e.g. an operational
    # child-only export), every row is treated as a child.
    has_row_type_col = any(v == 'row_type' for v in col_map.values())

    for i, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= header_row:
            continue
        if not raw_row or all(
            c is None or (isinstance(c, str) and not c.strip()) for c in raw_row
        ):
            continue

        row: dict[str, Any] = {}
        row_type_raw = ''
        for col_idx, field_name in col_map.items():
            if col_idx >= len(raw_row):
                continue
            value = raw_row[col_idx]
            if field_name == 'row_type':
                row_type_raw = _norm(value)
                continue
            if field_name in ('issue_date', 'movement_date'):
                row[field_name] = _to_date(value)
            elif field_name in DECIMAL_FIELDS:
                row[field_name] = _to_decimal(value)
            elif field_name == 'harvest_year':
                row[field_name] = _to_str(value)[:4]
            else:
                row[field_name] = _to_str(value)

        row_type = ROW_TYPE_MAP.get(row_type_raw, '')
        if not row_type and not has_row_type_col:
            row_type = ROW_TYPE_CHILD

        report.rows_total += 1

        # Quantity on child rows comes in negative; keep absolute value.
        qty = row.get('quantity_kg')
        if row_type == ROW_TYPE_CHILD and qty is not None:
            qty = abs(qty)
            row['quantity_kg'] = qty

        error = ''
        if not row_type:
            error = 'Tipo NF inválido. Use "Compra Futura" ou "Remessa".'
        elif not row.get('nf_number'):
            error = 'Nr. Nota obrigatório.'
        elif qty is None or qty <= 0:
            error = 'Quantidade deve ser diferente de zero.'
        elif not row.get('lot_number'):
            error = 'Contrato (Lote de Compra) obrigatório.'

        parsed = ParsedRow(
            row_number=i, row_type=row_type or '', data=row, error=error,
        )
        report.parsed.append(parsed)
        if error:
            report.rows_invalid += 1
        else:
            report.rows_valid += 1

    return report
