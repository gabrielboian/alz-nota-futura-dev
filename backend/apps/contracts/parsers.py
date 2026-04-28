"""XLSX parser for contract base uploads (BD-Baixa-lote-compra)."""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import IO, Any

import openpyxl

from .models import ContractBaseLot, ContractManagedLot, ContractUpload


# Column order must match BD-Baixa-lote-compra header row (row 3 in xlsx, 0-indexed 2).
COLUMN_ORDER = [
    'cpf_cnpj', 'liquidity', 'branch_name', 'producer_name', 'lot_type',
    'product', 'city', 'state_code', 'lot_number', 'quantity_kg',
    'delivered_kg', 'reversed_kg', 'remaining_kg', 'balance', 'payment_date',
    'price', 'freight_type', 'freight_value', 'emits_re', 'purchase_desk_id',
    'address_code', 'product_type', 'unit_value', 'lot_date', 'dap',
    'load_producer', 'load_location', 'load_city', 'load_state', 'cpf_cnpj_load',
    'delivery_start_date', 'delivery_end_date', 'destination_branch',
]

DATE_FIELDS = {'payment_date', 'lot_date', 'delivery_start_date', 'delivery_end_date'}
DECIMAL_FIELDS = {
    'quantity_kg', 'delivered_kg', 'reversed_kg', 'remaining_kg', 'balance',
    'price', 'freight_value', 'unit_value',
}
BOOLEAN_FIELDS = {'emits_re'}


def _to_str(v: Any) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == '':
        return Decimal('0')
    try:
        return Decimal(str(v).replace(',', '.'))
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _to_date(v: Any):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = _to_str(v).upper()
    return s in ('S', 'SIM', 'YES', 'TRUE', '1')


@dataclass
class UploadResult:
    upload: ContractUpload
    rows_created: int
    rows_updated: int
    rows_errored: int
    errors: list[str]


def parse_contract_xlsx(
    fp: IO[bytes], upload: ContractUpload, *, create_managed: bool = True
) -> UploadResult:
    """Parse xlsx and upsert `ContractBaseLot` (+ optional `ContractManagedLot`).

    Identity key: ``lot_number`` (first match wins). Existing rows are updated
    in place; new rows are created and receive a fresh `ContractManagedLot`.

    Header row is detected by looking for a row that contains 'CPF/CNPJ' in col 0.
    Data rows follow until a row where lot_number is empty.
    """
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.worksheets[0]
    # Prefer BD-Baixa-lote-compra if present; otherwise first sheet.
    for name in ('BD-Baixa-lote-compra', 'BD-Baixa-Lote-Compra'):
        if name in wb.sheetnames:
            ws = wb[name]
            break

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == 'CPF/CNPJ':
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(
            'Linha de cabeçalho com "CPF/CNPJ" não encontrada na primeira coluna — '
            'verifique se a aba "BD-Baixa-lote-compra" está presente.'
        )

    created = 0
    updated = 0
    errors: list[str] = []

    for i, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 2, values_only=True), start=header_row_idx + 2
    ):
        if not row or all(v is None for v in row):
            break
        lot_number = _to_str(row[8]) if len(row) > 8 else ''
        if not lot_number:
            break

        fields: dict[str, Any] = {}
        try:
            for idx, field in enumerate(COLUMN_ORDER):
                value = row[idx] if idx < len(row) else None
                if field in DATE_FIELDS:
                    fields[field] = _to_date(value)
                elif field in DECIMAL_FIELDS:
                    fields[field] = _to_decimal(value)
                elif field in BOOLEAN_FIELDS:
                    fields[field] = _to_bool(value)
                else:
                    fields[field] = _to_str(value)[:300]
        except Exception as exc:  # noqa: BLE001
            errors.append(f'Linha {i}: {exc}')
            continue

        existing = ContractBaseLot.objects.filter(lot_number=lot_number).first()
        if existing is None:
            base = ContractBaseLot.objects.create(upload=upload, **fields)
            created += 1
            if create_managed:
                ContractManagedLot.objects.get_or_create(
                    base_lot=base,
                    defaults={
                        'pickup_location': base.load_city or '',
                        'loading_site': base.load_location or '',
                        'collection_point_code': base.address_code or '',
                    },
                )
        else:
            changed = False
            for field_name, value in fields.items():
                if getattr(existing, field_name) != value:
                    setattr(existing, field_name, value)
                    changed = True
            # Always repoint to the most recent upload so admins can trace the
            # last touch, even when no field changed.
            if existing.upload_id != upload.pk:
                existing.upload = upload
                changed = True
            if changed:
                existing.save()
                updated += 1
            if create_managed:
                ContractManagedLot.objects.get_or_create(
                    base_lot=existing,
                    defaults={
                        'pickup_location': existing.load_city or '',
                        'loading_site': existing.load_location or '',
                        'collection_point_code': existing.address_code or '',
                    },
                )

    upload.row_count = created + updated
    upload.error_count = len(errors)
    upload.status = (
        ContractUpload.Status.SUCCESS if not errors else ContractUpload.Status.ERROR
    )
    if errors:
        upload.observations = '\n'.join(errors[:50])
    upload.save(update_fields=['row_count', 'error_count', 'status', 'observations'])

    return UploadResult(
        upload=upload,
        rows_created=created,
        rows_updated=updated,
        rows_errored=len(errors),
        errors=errors,
    )
