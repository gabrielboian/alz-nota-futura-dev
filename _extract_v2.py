"""Extract and diff V1 vs V2 xlsx files. Writes plain-text dumps per version."""
from pathlib import Path
import openpyxl
import json

ROOT = Path("/Users/gabrielboian/Desktop/dev.nosync/t2c/alz-nota-futura")


def dump_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None and str(v).strip() for v in row):
                rows.append([str(v) if v is not None else "" for v in row])
        result[sheet_name] = rows
    return result


def write_dump(data: dict, out_path: Path):
    lines = []
    for sheet_name, rows in data.items():
        lines.append(f"\n{'=' * 80}\nSHEET: {sheet_name}\n{'=' * 80}")
        for i, row in enumerate(rows, 1):
            lines.append(f"R{i}: " + " | ".join(row))
    out_path.write_text("\n".join(lines), encoding="utf-8")


V1_FORMS = ROOT / "v1" / "Formulários - NF Entrega Futura.xlsx"
V1_DB = ROOT / "v1" / "Rascunho Banco de Dados - NF Entrega Futura.xlsx"
V2_FORMS = ROOT / "v2" / "Formulários - NF Entrega Futura (1).xlsx"
V2_DB = ROOT / "v2" / "Rascunho Banco de Dados - NF Entrega Futura (1).xlsx"

v1_forms = dump_xlsx(V1_FORMS)
v1_db = dump_xlsx(V1_DB)
v2_forms = dump_xlsx(V2_FORMS)
v2_db = dump_xlsx(V2_DB)

write_dump(v1_forms, ROOT / "v1" / "_forms_dump.txt")
write_dump(v1_db, ROOT / "v1" / "_db_dump.txt")
write_dump(v2_forms, ROOT / "v2" / "_forms_dump.txt")
write_dump(v2_db, ROOT / "v2" / "_db_dump.txt")


def summary(name, data):
    print(f"\n=== {name} ===")
    for sheet, rows in data.items():
        ncols = max((len(r) for r in rows), default=0)
        print(f"  {sheet}: {len(rows)} rows, max {ncols} cols")


summary("V1 Forms", v1_forms)
summary("V2 Forms", v2_forms)
summary("V1 DB", v1_db)
summary("V2 DB", v2_db)


def sheet_diff(name, a, b):
    print(f"\n=== DIFF {name} ===")
    sa, sb = set(a.keys()), set(b.keys())
    added = sb - sa
    removed = sa - sb
    if added:
        print(f"  + added sheets: {sorted(added)}")
    if removed:
        print(f"  - removed sheets: {sorted(removed)}")
    for s in sa & sb:
        if a[s] != b[s]:
            print(f"  ~ changed sheet: {s} (V1 {len(a[s])} rows -> V2 {len(b[s])} rows)")


sheet_diff("Forms", v1_forms, v2_forms)
sheet_diff("DB", v1_db, v2_db)
print("\nDumps written to v1/_forms_dump.txt, v1/_db_dump.txt, v2/_forms_dump.txt, v2/_db_dump.txt")
