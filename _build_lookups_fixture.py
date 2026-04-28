"""Build core lookups fixture from V2 xlsx."""
import json
import os
import uuid
from collections import Counter
from datetime import datetime, timezone

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, 'v2/Rascunho Banco de Dados - NF Entrega Futura (1).xlsx')


def S(v):
    return str(v).strip() if v is not None else ''


def main():
    wb = openpyxl.load_workbook(XLSX)
    ts = datetime.now(timezone.utc).isoformat()
    out = []

    # Branches: row layout (None, None, state, sap_code, cnpj, description, type)
    ws = wb['BD-Filiais']
    seen_sap = set()
    type_map = {'Armazém': 'warehouse', 'Escritório': 'office'}
    for row in ws.iter_rows(values_only=True):
        if row[3] is None or S(row[3]) == 'Código Filiais':
            continue
        sap_code = S(row[3])
        if sap_code in seen_sap:
            continue
        seen_sap.add(sap_code)
        out.append({
            'model': 'core.branch',
            'pk': str(uuid.uuid4()),
            'fields': {
                'sap_code': sap_code,
                'state': S(row[2]),
                'cnpj': S(row[4]),
                'description': S(row[5]),
                'type': type_map.get(S(row[6]), 'warehouse'),
                'created_at': ts,
                'updated_at': ts,
            },
        })

    # Terminals: (None, name, sap_client, sap_supplier)
    ws = wb['BD-Terminais Destino']
    seen_name = set()
    for row in ws.iter_rows(values_only=True):
        name = S(row[1]) if len(row) > 1 else ''
        if not name or name == 'Nome Terminal' or name in seen_name:
            continue
        seen_name.add(name)
        out.append({
            'model': 'core.terminaldestination',
            'pk': str(uuid.uuid4()),
            'fields': {
                'name': name,
                'sap_client_code': S(row[2]) if len(row) > 2 else '',
                'sap_supplier_code': S(row[3]) if len(row) > 3 else '',
                'created_at': ts,
                'updated_at': ts,
            },
        })

    # Participants / Corredor / Comercial Responsável sheets contain only headers in V2.

    dest = os.path.join(ROOT, 'backend/apps/core/fixtures/lookups.json')
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    with open(dest, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f'Wrote {len(out)} records to {dest}')
    print('Per model:', dict(Counter(o['model'] for o in out)))


if __name__ == '__main__':
    main()
